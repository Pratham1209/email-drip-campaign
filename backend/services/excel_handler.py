
# import os
# import io
# from google.auth.transport.requests import Request
# from google.oauth2.credentials import Credentials  # Correct import for Credentials
# from google_auth_oauthlib.flow import InstalledAppFlow
# from googleapiclient.discovery import build
# from googleapiclient.http import MediaIoBaseUpload
# from openpyxl import Workbook

# # Google Drive API Scopes
# SCOPES = ['https://www.googleapis.com/auth/drive']

# # Path to token and credentials
# TOKEN_PATH = 'token.json'
# CREDENTIALS_PATH = 'credentials.json'  # Your OAuth 2.0 credentials

# # Your Google Drive File ID (The file you want to update)
# FILE_ID = '1ynbIxCcD-heVaIpSNUJADkDRZ-9yDKVX'  # Replace with your actual file ID from Google Drive

# # Function to authenticate Google Drive
# def authenticate_google_drive():
#     creds = None
#     # Check if token.json exists, which stores the user's access and refresh tokens
#     if os.path.exists(TOKEN_PATH):
#         creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
    
#     # If there are no (valid) credentials available, let the user log in.
#     if not creds or not creds.valid:
#         if creds and creds.expired and creds.refresh_token:
#             creds.refresh(Request())
#         else:
#             flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_PATH, SCOPES)
#             creds = flow.run_local_server(port=0)
        
#         # Save the credentials for the next run
#         with open(TOKEN_PATH, 'w') as token:
#             token.write(creds.to_json())

#     # Return the authenticated Google Drive service
#     return build('drive', 'v3', credentials=creds)

# # Function to update an Excel file on Google Drive
# def update_drive_file(service, file_id, new_data):
#     # Create a new Excel workbook or modify an existing one
#     wb = Workbook()
#     sheet = wb.active
#     sheet.append(["Account Name", "Contact Name", "Email Address", "Job Title", "Subject", "Body", "Call to Action"])

#     for account in new_data:
#         for email in account['emails']:
#             for contact in account['contacts']:
#                 sheet.append([
#                     account["account_name"],
#                     contact["name"],
#                     contact["email"],
#                     contact.get("job_title", ""),
#                     email["subject"],
#                     email["body"],
#                     email["call_to_action"]
#                 ])

#     # Save to a buffer (in memory) and upload to Google Drive
#     file_stream = io.BytesIO()
#     wb.save(file_stream)
#     file_stream.seek(0)

#     # Prepare the file for upload
#     media_body = MediaIoBaseUpload(file_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#     # Update the file on Google Drive
#     updated_file = service.files().update(fileId=file_id, media_body=media_body).execute()
#     return updated_file

# # Function to update the Google Drive Excel file
# def handle_post_request(new_data):
#     service = authenticate_google_drive()
#     updated_file = update_drive_file(service, FILE_ID, new_data)
#     print(f"Updated file: {updated_file}")
#     return updated_file

import os
import io
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials  # Correct import for Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from openpyxl import load_workbook, Workbook

# Google Drive API Scopes
SCOPES = ['https://www.googleapis.com/auth/drive']

# Path to token and credentials
TOKEN_PATH = 'token.json'
CREDENTIALS_PATH = 'credentials.json'  # Your OAuth 2.0 credentials

# Your Google Drive File ID (The file you want to update)
FILE_ID = '1ynbIxCcD-heVaIpSNUJADkDRZ-9yDKVX'  # Replace with your actual file ID from Google Drive

# Function to authenticate Google Drive
def authenticate_google_drive():
    creds = None
    # Check if token.json exists, which stores the user's access and refresh tokens
    if os.path.exists(TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
    
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_PATH, SCOPES)
            creds = flow.run_local_server(port=0)
        
        # Save the credentials for the next run
        with open(TOKEN_PATH, 'w') as token:
            token.write(creds.to_json())

    # Return the authenticated Google Drive service
    return build('drive', 'v3', credentials=creds)

# Function to download the file from Google Drive
def download_drive_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    file_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(file_stream, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    file_stream.seek(0)
    return file_stream

# Function to update an Excel file on Google Drive
def update_drive_file(service, file_id, new_data):
    try:
        # Step 1: Download the existing file
        file_stream = download_drive_file(service, file_id)

        # Step 2: Load the workbook and the active sheet
        try:
            wb = load_workbook(file_stream)
        except Exception:
            wb = Workbook()  # Create a new workbook if loading fails
        sheet = wb.active

        # Step 3: Append new data
        for account in new_data:
            for email in account['emails']:
                for contact in account['contacts']:
                    sheet.append([
                        account["account_name"],
                        contact["name"],
                        contact["email"],
                        contact.get("job_title", ""),
                        email["subject"],
                        email["body"],
                        email["call_to_action"]
                    ])

        # Step 4: Save to a buffer (in memory) and upload to Google Drive
        updated_file_stream = io.BytesIO()
        wb.save(updated_file_stream)
        updated_file_stream.seek(0)

        # Prepare the file for upload
        media_body = MediaIoBaseUpload(updated_file_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Step 5: Update the file on Google Drive
        updated_file = service.files().update(fileId=file_id, media_body=media_body).execute()
        return updated_file

    except Exception as e:
        print(f"An error occurred: {e}")
        raise

# Function to update the Google Drive Excel file
def handle_post_request(new_data):
    service = authenticate_google_drive()
    updated_file = update_drive_file(service, FILE_ID, new_data)
    print(f"Updated file: {updated_file}")
    return updated_file
